import argparse
import calendar
import re
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook


HEADER_ALIASES = {
    "contract_id": ["contract_id", "合同编号", "合同号", "合同编码"],
    "lease_start": ["lease_start", "起租日", "开始日期", "起始日期", "租赁开始"],
    "lease_end": ["lease_end", "止租日", "结束日期", "终止日期", "到期日", "租赁结束"],
    "payment_amount": ["payment_amount", "租金", "付款金额", "支付金额", "每期租金", "租赁付款"],
    "payment_frequency": ["payment_frequency", "付款频率", "支付频率", "频率"],
    "discount_rate": ["discount_rate", "折现率", "贴现率", "年利率"],
    "payment_timing": ["payment_timing", "付款时点", "期初期末", "期初/期末"],
    "currency": ["currency", "币种", "币别"],
}


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[()（）\[\]【】:%/\\-]", "", text)
    return text


def find_header_index(headers, aliases):
    alias_norms = [normalize_header(a) for a in aliases]
    for alias in alias_norms:
        if alias in headers:
            return headers.index(alias)
    for idx, header in enumerate(headers):
        for alias in alias_norms:
            if alias and alias in header:
                return idx
    return None


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date: {value}")


def add_months(d, months):
    year = d.year + (d.month - 1 + months) // 12
    month = (d.month - 1 + months) % 12 + 1
    day = min(d.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def frequency_to_months(value):
    text = str(value).strip().lower()
    if text in ("m", "month", "monthly"):
        return 1
    if text in ("q", "quarter", "quarterly"):
        return 3
    if text in ("a", "y", "annual", "year", "yearly"):
        return 12
    if text in ("月", "月度", "每月"):
        return 1
    if text in ("季", "季度", "每季"):
        return 3
    if text in ("年", "年度", "每年"):
        return 12
    raise ValueError(f"Invalid payment_frequency: {value}")


def parse_rate(value):
    num = parse_number(value)
    if num is None:
        raise ValueError(f"Invalid discount_rate: {value}")
    if num > 1:
        num = num / 100.0
    return num


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return None


def read_leases(ws):
    required = [
        "contract_id",
        "lease_start",
        "lease_end",
        "payment_amount",
        "payment_frequency",
        "discount_rate",
        "payment_timing",
    ]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(value) for value in header_row]
    indices = {}
    for key in required:
        idx = find_header_index(headers, HEADER_ALIASES.get(key, [key]))
        if idx is None:
            raise ValueError(f"Missing required column: {key}")
        indices[key] = idx
    currency_idx = find_header_index(headers, HEADER_ALIASES["currency"])

    leases = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        lease = {}
        for key in required:
            lease[key] = row[indices[key]]
        lease["currency"] = row[currency_idx] if currency_idx is not None else ""
        if not lease["contract_id"]:
            raise ValueError(f"Row {row_idx} missing contract_id.")
        leases.append(lease)
    return leases


def generate_payment_dates(start, end, freq_months, timing):
    dates = []
    if timing == "begin":
        current = start
    else:
        current = add_months(start, freq_months)
    while current <= end:
        dates.append(current)
        current = add_months(current, freq_months)
    return dates


def calculate_schedule(lease):
    start = parse_date(lease["lease_start"])
    end = parse_date(lease["lease_end"])
    if not start or not end or end < start:
        raise ValueError(f"Invalid lease dates for {lease['contract_id']}")

    payment = parse_number(lease["payment_amount"])
    if payment is None:
        raise ValueError(f"Invalid payment_amount for {lease['contract_id']}")
    freq_months = frequency_to_months(lease["payment_frequency"])
    annual_rate = parse_rate(lease["discount_rate"])
    timing = str(lease["payment_timing"]).strip().lower()
    if timing in ("期初", "期初付款", "期初付", "月初"):
        timing = "begin"
    elif timing in ("期末", "期末付款", "期末付", "月末"):
        timing = "end"
    if timing not in ("begin", "end"):
        raise ValueError(f"Invalid payment_timing for {lease['contract_id']}: {timing}")

    dates = generate_payment_dates(start, end, freq_months, timing)
    if not dates:
        raise ValueError(f"No payment dates for {lease['contract_id']}")

    periods_per_year = 12 / freq_months
    periodic_rate = annual_rate / periods_per_year if periods_per_year else 0.0
    n = len(dates)

    if periodic_rate == 0:
        pv = payment * n
    else:
        pv = payment * (1 - (1 + periodic_rate) ** (-n)) / periodic_rate
        if timing == "begin":
            pv = pv * (1 + periodic_rate)

    schedule = []
    opening = pv
    total_interest = 0.0
    for idx, pay_date in enumerate(dates, start=1):
        if timing == "begin":
            principal = payment
            balance_after_payment = opening - principal
            interest = balance_after_payment * periodic_rate
            closing = balance_after_payment + interest
        else:
            interest = opening * periodic_rate
            principal = payment - interest
            closing = opening - principal

        schedule.append(
            {
                "period": idx,
                "payment_date": pay_date,
                "opening_balance": opening,
                "payment": payment,
                "interest": interest,
                "principal": principal,
                "closing_balance": closing,
            }
        )
        total_interest += interest
        opening = closing

    return pv, total_interest, schedule


def write_output(path, results):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(
        [
            "contract_id",
            "currency",
            "payment_amount",
            "periods",
            "annual_rate",
            "initial_liability",
            "total_interest",
            "ending_balance",
        ]
    )

    ws_schedule = wb.create_sheet("Schedule")
    ws_schedule.append(
        [
            "contract_id",
            "period",
            "payment_date",
            "opening_balance",
            "payment",
            "interest",
            "principal",
            "closing_balance",
        ]
    )

    for result in results:
        lease = result["lease"]
        pv = result["pv"]
        total_interest = result["total_interest"]
        schedule = result["schedule"]
        ending_balance = schedule[-1]["closing_balance"] if schedule else 0
        ws_summary.append(
            [
                lease["contract_id"],
                lease.get("currency", ""),
                float(lease["payment_amount"]),
                len(schedule),
                parse_rate(lease["discount_rate"]),
                pv,
                total_interest,
                ending_balance,
            ]
        )
        for row in schedule:
            ws_schedule.append(
                [
                    lease["contract_id"],
                    row["period"],
                    row["payment_date"],
                    row["opening_balance"],
                    row["payment"],
                    row["interest"],
                    row["principal"],
                    row["closing_balance"],
                ]
            )

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Lease amortization schedule generator.")
    parser.add_argument("--input", default="input.xlsx", help="Input .xlsx (default: input.xlsx).")
    parser.add_argument("--output", default="output.xlsx", help="Output .xlsx (default: output.xlsx).")
    parser.add_argument("--sheet", default="Leases", help="Sheet name (default: Leases).")
    args = parser.parse_args()

    try:
        wb = load_workbook(args.input, data_only=True)
        ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
        leases = read_leases(ws)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    results = []
    try:
        for lease in leases:
            pv, total_interest, schedule = calculate_schedule(lease)
            results.append(
                {"lease": lease, "pv": pv, "total_interest": total_interest, "schedule": schedule}
            )
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    write_output(Path(args.output), results)
    print(f"Saved output: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
