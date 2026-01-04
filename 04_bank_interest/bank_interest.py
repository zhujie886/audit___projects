import argparse
import re
import sys
from datetime import date, datetime

from openpyxl import Workbook, load_workbook


COLUMN_ALIASES = {
    "principal": ["principal", "本金", "存款本金", "存款金额", "金额", "余额", "期末余额", "本币余额"],
    "annual_rate": ["annual_rate", "年利率", "利率", "年化利率", "执行利率"],
    "days": ["days", "天数", "计息天数", "期限天数", "存期天数", "存期"],
    "start_date": ["start_date", "起息日", "开始日期", "起始日期", "起息日期"],
    "end_date": ["end_date", "到期日", "结束日期", "终止日期", "到期日期"],
    "day_count": ["day_count", "计息基数", "年天数", "天数基数"],
    "account": ["account", "账号", "账户", "银行账号", "账户号"],
    "direction": ["direction", "余额方向", "借贷方向", "方向"],
}


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[()（）\[\]【】:%/\\-]", "", text)
    return text


def find_header_index(headers, aliases):
    alias_norms = [normalize_header(a) for a in aliases]
    for alias in alias_norms:
        if alias in headers:
            return headers.index(alias)
    for idx, header in enumerate(headers):
        for alias in alias_norms:
            if alias and alias in header:
                return idx
    return None


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Invalid date: {value}")


def parse_rate(value):
    num = parse_number(value)
    if num is None:
        raise ValueError(f"Invalid annual_rate: {value}")
    if num > 1:
        num = num / 100.0
    return num


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return None


def read_rows(ws):
    required = ["principal", "annual_rate"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(value) for value in header_row]
    indices = {}
    for key in required:
        idx = find_header_index(headers, COLUMN_ALIASES[key])
        if idx is None:
            raise ValueError(f"Missing required column: {key}")
        indices[key] = idx
    for key in ("days", "start_date", "end_date", "day_count", "account", "direction"):
        indices[key] = find_header_index(headers, COLUMN_ALIASES[key])
    return indices


def main():
    parser = argparse.ArgumentParser(description="Bank deposit interest calculator.")
    parser.add_argument("--input", default="input.xlsx", help="Input .xlsx (default: input.xlsx).")
    parser.add_argument("--output", default="output.xlsx", help="Output .xlsx (default: output.xlsx).")
    parser.add_argument("--sheet", default="Deposits", help="Sheet name (default: Deposits).")
    args = parser.parse_args()

    try:
        wb = load_workbook(args.input, data_only=True)
        ws = wb[args.sheet] if args.sheet in wb.sheetnames else wb.active
        headers = read_rows(ws)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = ws.title

    header_values = [cell.value for cell in ws[1]]
    ws_out.append(header_values + ["days_calc", "interest", "maturity_amount"])

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        principal = row[headers["principal"]]
        annual_rate = row[headers["annual_rate"]]
        if principal is None or annual_rate is None:
            raise ValueError(f"Row {row_idx} missing principal or annual_rate.")

        principal = parse_number(principal)
        if principal is None:
            raise ValueError(f"Row {row_idx} invalid principal.")
        annual_rate = parse_rate(annual_rate)

        days_val = row[headers.get("days")] if headers.get("days") is not None else None
        start_val = row[headers.get("start_date")] if headers.get("start_date") is not None else None
        end_val = row[headers.get("end_date")] if headers.get("end_date") is not None else None
        day_count = row[headers.get("day_count")] if headers.get("day_count") is not None else None
        day_count = int(day_count) if day_count else 365

        if days_val is not None:
            days = int(days_val)
        else:
            start_date = parse_date(start_val)
            end_date = parse_date(end_val)
            if not start_date or not end_date:
                raise ValueError(
                    f"Row {row_idx} requires days or start_date/end_date."
                )
            days = (end_date - start_date).days

        interest = principal * annual_rate * days / day_count
        maturity = principal + interest

        ws_out.append(list(row) + [days, interest, maturity])

    wb_out.save(args.output)
    print(f"Saved output: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
