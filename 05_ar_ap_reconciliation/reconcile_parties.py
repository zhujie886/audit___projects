import argparse
import re
import sys
from collections import defaultdict

from openpyxl import Workbook, load_workbook


PARTY_HEADERS = [
    "party",
    "counterparty",
    "vendor",
    "customer",
    "name",
    "supplier",
    "往来单位",
    "单位名称",
    "客户名称",
    "供应商名称",
    "对方单位",
    "客商名称",
]
AMOUNT_HEADERS = ["amount", "balance", "value", "total", "余额", "期末余额", "本币余额", "金额", "发生额"]
DEBIT_HEADERS = ["debit", "借方余额", "借方发生额", "期末借方余额"]
CREDIT_HEADERS = ["credit", "贷方余额", "贷方发生额", "期末贷方余额"]
ACCOUNT_HEADERS = ["account_name", "account", "科目名称", "科目", "会计科目"]
CODE_HEADERS = ["account_code", "account", "code", "科目编码", "科目代码", "科目编号"]
DIRECTION_HEADERS = ["direction", "余额方向", "借贷方向", "方向"]

SHEET_ALIASES = {
    "AR": ["AR", "应收账款", "应收", "客户应收"],
    "AP": ["AP", "应付账款", "应付", "供应商应付"],
    "OtherAR": ["OtherAR", "其他应收", "其他应收款"],
    "OtherAP": ["OtherAP", "其他应付", "其他应付款"],
}


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[()（）\[\]【】:%/\\-]", "", text)
    return text


def find_header_index(headers, options):
    option_norms = [normalize_header(o) for o in options]
    for option in option_norms:
        if option in headers:
            return headers.index(option)
    for idx, header in enumerate(headers):
        for option in option_norms:
            if option and option in header:
                return idx
    return None


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return None


def direction_sign(value):
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in ("贷", "贷方", "credit", "c"):
        return -1
    if text in ("借", "借方", "debit", "d"):
        return 1
    return None


def classify_account(account_name, account_code):
    name = str(account_name or "").strip()
    code = str(account_code or "").strip()
    if "应收账款" in name:
        return "AR"
    if "应付账款" in name:
        return "AP"
    if "其他应收" in name:
        return "OtherAR"
    if "其他应付" in name:
        return "OtherAP"
    if code.startswith("1122"):
        return "AR"
    if code.startswith("2202"):
        return "AP"
    if code.startswith("1221"):
        return "OtherAR"
    if code.startswith("2241"):
        return "OtherAP"
    return None


def read_sheet(ws, category=None):
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(value) for value in header_row]
    party_idx = find_header_index(headers, PARTY_HEADERS)
    amount_idx = find_header_index(headers, AMOUNT_HEADERS)
    debit_idx = find_header_index(headers, DEBIT_HEADERS)
    credit_idx = find_header_index(headers, CREDIT_HEADERS)
    account_idx = find_header_index(headers, ACCOUNT_HEADERS)
    code_idx = find_header_index(headers, CODE_HEADERS)
    direction_idx = find_header_index(headers, DIRECTION_HEADERS)

    if party_idx is None:
        raise ValueError(f"Missing party column in sheet {ws.title}")
    if amount_idx is None and (debit_idx is None or credit_idx is None):
        raise ValueError(f"Missing amount columns in sheet {ws.title}")

    if category:
        totals_by_category = {category: defaultdict(float)}
    else:
        totals_by_category = {
            "AR": defaultdict(float),
            "AP": defaultdict(float),
            "OtherAR": defaultdict(float),
            "OtherAP": defaultdict(float),
        }
    unclassified = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        party = row[party_idx]
        if party is None:
            continue
        party_key = str(party).strip()
        if not party_key:
            continue

        amount = None
        if amount_idx is not None:
            amount = parse_number(row[amount_idx])
        if amount is None and debit_idx is not None and credit_idx is not None:
            debit = parse_number(row[debit_idx]) or 0.0
            credit = parse_number(row[credit_idx]) or 0.0
            amount = debit - credit
        if amount is None:
            continue

        if direction_idx is not None and amount >= 0:
            sign = direction_sign(row[direction_idx])
            if sign is not None:
                amount = amount * sign

        if category:
            totals_by_category[category][party_key] += amount
        else:
            account_name = row[account_idx] if account_idx is not None else ""
            account_code = row[code_idx] if code_idx is not None else ""
            cat = classify_account(account_name, account_code)
            if not cat:
                unclassified.append((party_key, account_code, account_name, amount))
                continue
            totals_by_category[cat][party_key] += amount

    return totals_by_category, unclassified


def main():
    parser = argparse.ArgumentParser(description="AR/AP reconciliation by party.")
    parser.add_argument("--input", default="input.xlsx", help="Input .xlsx (default: input.xlsx).")
    parser.add_argument("--output", default="output.xlsx", help="Output .xlsx (default: output.xlsx).")
    parser.add_argument("--ar-sheet", default="AR", help="AR sheet name.")
    parser.add_argument("--ap-sheet", default="AP", help="AP sheet name.")
    parser.add_argument("--other-ar-sheet", default="OtherAR", help="Other AR sheet name.")
    parser.add_argument("--other-ap-sheet", default="OtherAP", help="Other AP sheet name.")
    args = parser.parse_args()

    try:
        wb = load_workbook(args.input, data_only=True)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    def normalize_sheet_name(name):
        return normalize_header(name)

    def pick_sheet_map():
        sheet_map = {}
        normalized = {name: normalize_sheet_name(name) for name in wb.sheetnames}
        for category, aliases in SHEET_ALIASES.items():
            alias_norms = [normalize_sheet_name(a) for a in aliases]
            for sheet, norm in normalized.items():
                if sheet in sheet_map:
                    continue
                for alias in alias_norms:
                    if alias and (norm == alias or alias in norm):
                        sheet_map[sheet] = category
                        break
        return sheet_map

    try:
        sheet_map = pick_sheet_map()
        unclassified_rows = []
        if sheet_map:
            totals = {
                "AR": defaultdict(float),
                "AP": defaultdict(float),
                "OtherAR": defaultdict(float),
                "OtherAP": defaultdict(float),
            }
            for sheet, category in sheet_map.items():
                sheet_totals, unclassified = read_sheet(wb[sheet], category=category)
                for key, values in sheet_totals.items():
                    for party, amount in values.items():
                        totals[key][party] += amount
                unclassified_rows.extend(unclassified)
        else:
            totals, unclassified_rows = read_sheet(wb.active, category=None)
        ar = totals.get("AR", defaultdict(float))
        ap = totals.get("AP", defaultdict(float))
        other_ar = totals.get("OtherAR", defaultdict(float))
        other_ap = totals.get("OtherAP", defaultdict(float))
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    parties = set(ar) | set(ap) | set(other_ar) | set(other_ap)

    wb_out = Workbook()
    ws_summary = wb_out.active
    ws_summary.title = "Summary"
    ws_summary.append(
        [
            "party",
            "AR",
            "AP",
            "OtherAR",
            "OtherAP",
            "receivable_total",
            "payable_total",
            "net_receivable",
            "has_both",
        ]
    )

    ws_issues = wb_out.create_sheet("Issues")
    ws_issues.append(
        ["party", "receivable_total", "payable_total", "net_receivable"]
    )

    ws_unclassified = wb_out.create_sheet("Unclassified")
    ws_unclassified.append(["party", "account_code", "account_name", "amount"])

    for party in sorted(parties):
        ar_val = ar.get(party, 0.0)
        ap_val = ap.get(party, 0.0)
        other_ar_val = other_ar.get(party, 0.0)
        other_ap_val = other_ap.get(party, 0.0)
        receivable = ar_val + other_ar_val
        payable = ap_val + other_ap_val
        net = receivable - payable
        has_both = "Y" if receivable > 0 and payable > 0 else "N"

        ws_summary.append(
            [
                party,
                ar_val,
                ap_val,
                other_ar_val,
                other_ap_val,
                receivable,
                payable,
                net,
                has_both,
            ]
        )

        if has_both == "Y":
            ws_issues.append([party, receivable, payable, net])

    for row in unclassified_rows:
        ws_unclassified.append(list(row))

    wb_out.save(args.output)
    print(f"Saved output: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
