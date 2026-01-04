import argparse
import sys
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def resolve_columns(ws, header_row, columns_arg):
    if not columns_arg:
        return None
    parts = [p.strip() for p in columns_arg.split(",") if p.strip()]
    columns = set()
    header_map = {
        normalize_header(ws.cell(row=header_row, column=col).value): col
        for col in range(1, ws.max_column + 1)
    }
    for part in parts:
        if part.isalpha():
            columns.add(column_index_from_string(part.upper()))
        else:
            key = normalize_header(part)
            if key not in header_map:
                raise ValueError(f"Unknown column: {part}")
            columns.add(header_map[key])
    return columns


def round_sheet(ws, header_row, decimals, columns):
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            if columns and cell.column not in columns:
                continue
            if cell.data_type == "f":
                continue
            value = cell.value
            if isinstance(value, (int, float)):
                if decimals == 0:
                    cell.value = int(round(value))
                else:
                    cell.value = round(value, decimals)
            if isinstance(value, (datetime, date)):
                continue


def main():
    parser = argparse.ArgumentParser(description="Round numeric values in Excel.")
    parser.add_argument("--input", default="input.xlsx", help="Input .xlsx (default: input.xlsx).")
    parser.add_argument("--output", default="output.xlsx", help="Output .xlsx (default: output.xlsx).")
    parser.add_argument("--decimals", type=int, default=2, help="Decimal places.")
    parser.add_argument("--sheets", default="", help="Comma-separated sheet names.")
    parser.add_argument("--columns", default="", help="Column letters or header names.")
    parser.add_argument("--header-row", type=int, default=1, help="Header row (default: 1).")
    args = parser.parse_args()

    try:
        wb = load_workbook(args.input)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    sheet_names = (
        [s.strip() for s in args.sheets.split(",") if s.strip()]
        if args.sheets
        else wb.sheetnames
    )

    for name in sheet_names:
        if name not in wb.sheetnames:
            print(f"ERROR: Sheet not found: {name}", file=sys.stderr)
            return 1
        ws = wb[name]
        try:
            columns = resolve_columns(ws, args.header_row, args.columns) if args.columns else None
        except ValueError as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1
        round_sheet(ws, args.header_row, args.decimals, columns)

    wb.save(args.output)
    print(f"Saved output: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
