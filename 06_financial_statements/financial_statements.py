import argparse
import re
import sys
from collections import defaultdict

from openpyxl import Workbook, load_workbook


TB_SHEET_CANDIDATES = [
    "TB_Current",
    "TrialBalance",
    "Trial Balance",
    "TB",
    "科目余额表",
    "试算平衡表",
    "余额表",
    "总账余额表",
]

MAPPING_SHEET_CANDIDATES = [
    "Mapping",
    "科目映射",
    "报表映射",
    "报表项目映射",
    "映射",
]

PARAMETERS_SHEET_CANDIDATES = ["Parameters", "参数", "设置"]

CODE_HEADERS = ["account_code", "account", "code", "科目编码", "科目代码", "科目编号"]
NAME_HEADERS = ["account_name", "name", "description", "科目名称", "科目", "科目全名"]
TYPE_HEADERS = ["account_type", "科目类型", "科目类别", "科目性质"]

END_BALANCE_HEADERS = ["ending_balance", "期末余额", "余额", "本币余额", "期末余额本币"]
END_DEBIT_HEADERS = ["期末借方余额", "期末借方余额本币", "期末借方", "借方余额", "借方期末余额"]
END_CREDIT_HEADERS = ["期末贷方余额", "期末贷方余额本币", "期末贷方", "贷方余额", "贷方期末余额"]

BEGIN_BALANCE_HEADERS = ["beginning_balance", "期初余额", "期初余额本币"]
BEGIN_DEBIT_HEADERS = ["期初借方余额", "期初借方余额本币", "期初借方", "借方期初余额"]
BEGIN_CREDIT_HEADERS = ["期初贷方余额", "期初贷方余额本币", "期初贷方", "贷方期初余额"]

DIRECTION_HEADERS = ["余额方向", "借贷方向", "方向", "方向借贷"]

MAPPING_STATEMENT_HEADERS = ["statement", "报表", "报表类型", "表"]
MAPPING_SECTION_HEADERS = ["section", "板块", "分类", "项目分类", "报表项目分类"]
MAPPING_LINE_HEADERS = ["line_item", "line", "item", "项目", "行项目", "报表项目", "项目名称"]
MAPPING_CODE_HEADERS = ["account_code", "account", "code", "科目编码", "科目代码", "科目编号"]
MAPPING_SIGN_HEADERS = ["sign", "符号", "系数", "正负"]

PARAM_KEY_ALIASES = {
    "cash_begin": ["cash_begin", "期初现金", "现金期初", "期初货币资金", "期初现金及现金等价物"],
    "cash_end": ["cash_end", "期末现金", "现金期末", "期末货币资金", "期末现金及现金等价物"],
    "tolerance": ["tolerance", "容差", "允许误差"],
}

REVENUE_KEYWORDS = [
    "收入",
    "主营业务收入",
    "其他业务收入",
    "投资收益",
    "公允价值变动收益",
    "资产处置收益",
    "营业外收入",
    "利息收入",
    "手续费收入",
]

EXPENSE_KEYWORDS = [
    "成本",
    "费用",
    "税金",
    "附加",
    "所得税",
    "损失",
    "营业外支出",
    "管理费用",
    "销售费用",
    "财务费用",
    "研发费用",
    "信用减值损失",
    "资产减值损失",
    "手续费支出",
]

CASH_KEYWORDS = ["现金", "银行存款", "库存现金", "货币资金", "现金等价物"]
CASH_CODE_PREFIXES = ("1001", "1002", "1012")


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[()（）\[\]【】:%/\\-]", "", text)
    return text


def find_header_index(headers, options):
    option_norms = [normalize_header(o) for o in options]
    for option in option_norms:
        if option in headers:
            return headers.index(option)
    for idx, header in enumerate(headers):
        for option in option_norms:
            if option and option in header:
                return idx
    return None


def parse_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return None


def direction_sign(value):
    if value is None:
        return None
    text = str(value).strip().lower()
    if text in ("贷", "贷方", "credit", "c"):
        return -1
    if text in ("借", "借方", "debit", "d"):
        return 1
    return None


def code_to_str(value):
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def pick_sheet(wb, candidates):
    normalized = {name: normalize_header(name) for name in wb.sheetnames}
    for candidate in candidates:
        cand_norm = normalize_header(candidate)
        for name, norm in normalized.items():
            if norm == cand_norm or cand_norm in norm:
                return name
    return None


def sheet_has_tb_headers(ws):
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(value) for value in header_row]
    code_idx = find_header_index(headers, CODE_HEADERS)
    bal_idx = find_header_index(headers, END_BALANCE_HEADERS)
    debit_idx = find_header_index(headers, END_DEBIT_HEADERS)
    credit_idx = find_header_index(headers, END_CREDIT_HEADERS)
    return code_idx is not None and (bal_idx is not None or (debit_idx is not None and credit_idx is not None))


def pick_tb_sheet(wb):
    sheet = pick_sheet(wb, TB_SHEET_CANDIDATES)
    if sheet:
        return sheet
    for name in wb.sheetnames:
        if sheet_has_tb_headers(wb[name]):
            return name
    return wb.sheetnames[0]


def compute_balance(row, balance_idx, debit_idx, credit_idx, direction_idx):
    if debit_idx is not None and credit_idx is not None:
        debit = parse_number(row[debit_idx]) or 0.0
        credit = parse_number(row[credit_idx]) or 0.0
        return debit - credit
    if balance_idx is None:
        return None
    value = parse_number(row[balance_idx])
    if value is None:
        return None
    if direction_idx is not None:
        sign = direction_sign(row[direction_idx])
        if sign is not None and value >= 0:
            value *= sign
    return value


def read_tb(ws):
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(value) for value in header_row]
    code_idx = find_header_index(headers, CODE_HEADERS)
    name_idx = find_header_index(headers, NAME_HEADERS)
    type_idx = find_header_index(headers, TYPE_HEADERS)
    end_bal_idx = find_header_index(headers, END_BALANCE_HEADERS)
    end_debit_idx = find_header_index(headers, END_DEBIT_HEADERS)
    end_credit_idx = find_header_index(headers, END_CREDIT_HEADERS)
    begin_bal_idx = find_header_index(headers, BEGIN_BALANCE_HEADERS)
    begin_debit_idx = find_header_index(headers, BEGIN_DEBIT_HEADERS)
    begin_credit_idx = find_header_index(headers, BEGIN_CREDIT_HEADERS)
    direction_idx = find_header_index(headers, DIRECTION_HEADERS)

    if code_idx is None:
        raise ValueError(f"Missing account_code column in {ws.title}")
    if end_bal_idx is None and (end_debit_idx is None or end_credit_idx is None):
        raise ValueError(f"Missing ending balance columns in {ws.title}")

    accounts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        code = code_to_str(row[code_idx])
        if not code:
            continue
        name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else ""
        acc_type = str(row[type_idx]).strip() if type_idx is not None and row[type_idx] else ""

        end_balance = compute_balance(row, end_bal_idx, end_debit_idx, end_credit_idx, direction_idx)
        begin_balance = compute_balance(row, begin_bal_idx, begin_debit_idx, begin_credit_idx, direction_idx)
        if end_balance is None:
            continue

        accounts.append(
            {
                "code": code,
                "name": name,
                "type": acc_type,
                "end_balance": end_balance,
                "begin_balance": begin_balance,
            }
        )
    return accounts


def build_tb_dict(accounts):
    balances = defaultdict(float)
    names = {}
    for acc in accounts:
        balance = acc.get("end_balance")
        if balance is None:
            continue
        balances[acc["code"]] += balance
        if acc.get("name"):
            names[acc["code"]] = acc["name"]
    return balances, names


def parse_code_tokens(value):
    text = code_to_str(value)
    if not text:
        return []
    tokens = re.split(r"[;,，、\s]+", text)
    return [t.strip() for t in tokens if t.strip()]


def normalize_statement(value):
    text = str(value).strip().lower()
    if text in ("bs", "balance sheet", "资产负债表"):
        return "BS"
    if text in ("is", "income statement", "利润表", "损益表"):
        return "IS"
    if text in ("cf", "cash flow", "cashflow", "现金流量表"):
        return "CF"
    return text.upper()


def iter_matching_codes(token, tb_balances):
    token = token.strip()
    if not token:
        return []
    if token.endswith("*"):
        prefix = token[:-1]
        return [code for code in tb_balances if code.startswith(prefix)]
    if "-" in token:
        start, end = token.split("-", 1)
        start = start.strip()
        end = end.strip()
        if start.isdigit() and end.isdigit():
            start_num = int(start)
            end_num = int(end)
            matches = []
            for code in tb_balances:
                if code.isdigit():
                    code_num = int(code)
                    if start_num <= code_num <= end_num:
                        matches.append(code)
            return matches
    if token in tb_balances:
        return [token]
    return []


def read_mapping(ws):
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(value) for value in header_row]
    stmt_idx = find_header_index(headers, MAPPING_STATEMENT_HEADERS)
    sec_idx = find_header_index(headers, MAPPING_SECTION_HEADERS)
    line_idx = find_header_index(headers, MAPPING_LINE_HEADERS)
    code_idx = find_header_index(headers, MAPPING_CODE_HEADERS)
    sign_idx = find_header_index(headers, MAPPING_SIGN_HEADERS)
    if None in (stmt_idx, sec_idx, line_idx, code_idx):
        raise ValueError(f"Missing required mapping columns in {ws.title}")

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue
        statement = normalize_statement(row[stmt_idx]) if row[stmt_idx] else ""
        section = str(row[sec_idx]).strip() if row[sec_idx] else ""
        line_item = str(row[line_idx]).strip() if row[line_idx] else ""
        code_cell = row[code_idx]
        sign = parse_number(row[sign_idx]) if sign_idx is not None else 1.0
        sign = sign if sign is not None else 1.0
        if not statement or not section or not line_item or code_cell is None:
            raise ValueError(f"Row {row_idx} missing mapping values.")
        codes = parse_code_tokens(code_cell)
        if not codes:
            raise ValueError(f"Row {row_idx} has empty account_code.")
        rows.append(
            {
                "statement": statement,
                "section": section,
                "line_item": line_item,
                "codes": codes,
                "sign": sign,
            }
        )
    return rows


def read_parameters(wb, sheet_name):
    if not sheet_name or sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    params = {}
    alias_map = {}
    for key, aliases in PARAM_KEY_ALIASES.items():
        for alias in aliases:
            alias_map[normalize_header(alias)] = key

    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or row[0] is None:
            continue
        raw_key = normalize_header(row[0])
        key = alias_map.get(raw_key, raw_key)
        params[key] = row[1] if len(row) > 1 else None
    return params


def section_match(section, keywords):
    text = str(section).strip().lower()
    return any(k in text for k in keywords)


def sum_sections(section_totals, predicate):
    total = 0.0
    found = False
    for section, amount in section_totals.items():
        if predicate(section):
            total += amount
            found = True
    return total, found


def classify_bs(account_type, code):
    acc_type = str(account_type or "")
    if "资产" in acc_type:
        return "Assets"
    if "负债" in acc_type:
        return "Liabilities"
    if "权益" in acc_type or "所有者权益" in acc_type:
        return "Equity"
    if code.startswith("1"):
        return "Assets"
    if code.startswith("2"):
        return "Liabilities"
    if code.startswith("3"):
        return "Equity"
    return None


def classify_is(account_type, account_name, code):
    name = str(account_name or "")
    acc_type = str(account_type or "")
    for keyword in REVENUE_KEYWORDS:
        if keyword in name:
            return "Revenue"
    for keyword in EXPENSE_KEYWORDS:
        if keyword in name:
            return "Expense"
    if "收入" in acc_type:
        return "Revenue"
    if "费用" in acc_type or "成本" in acc_type:
        return "Expense"
    if "损益" in acc_type:
        if code.startswith("6"):
            return "Revenue"
        if code.startswith("5") or code.startswith("4"):
            return "Expense"
    if code.startswith("6"):
        return "Revenue"
    if code.startswith("5") or code.startswith("4"):
        return "Expense"
    return None


def is_cash_account(account):
    name = account.get("name", "")
    code = account.get("code", "")
    if any(keyword in name for keyword in CASH_KEYWORDS):
        return True
    return code.startswith(CASH_CODE_PREFIXES)


def compute_cash_totals(accounts):
    cash_begin_values = []
    cash_end_values = []
    for acc in accounts:
        if not is_cash_account(acc):
            continue
        if acc.get("begin_balance") is not None:
            cash_begin_values.append(acc["begin_balance"])
        if acc.get("end_balance") is not None:
            cash_end_values.append(acc["end_balance"])
    cash_begin = sum(cash_begin_values) if cash_begin_values else None
    cash_end = sum(cash_end_values) if cash_end_values else None
    return cash_begin, cash_end


def apply_mapping(tb_balances, mapping_rows):
    section_order = defaultdict(list)
    line_order = defaultdict(lambda: defaultdict(list))
    line_totals = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    section_totals = defaultdict(lambda: defaultdict(float))
    used_accounts = set()
    missing_accounts = set()

    for row in mapping_rows:
        statement = row["statement"]
        section = row["section"]
        line_item = row["line_item"]
        if section not in section_order[statement]:
            section_order[statement].append(section)
        if line_item not in line_order[statement][section]:
            line_order[statement][section].append(line_item)

        for token in row["codes"]:
            matched_codes = iter_matching_codes(token, tb_balances)
            if not matched_codes:
                missing_accounts.add(token)
                continue
            for code in matched_codes:
                amount = tb_balances[code] * row["sign"]
                line_totals[statement][section][line_item] += amount
                section_totals[statement][section] += amount
                used_accounts.add(code)

    return section_order, line_order, line_totals, section_totals, used_accounts, missing_accounts


def main():
    parser = argparse.ArgumentParser(description="Financial statements generator with checks.")
    parser.add_argument("--input", default="input.xlsx", help="Input .xlsx (default: input.xlsx).")
    parser.add_argument("--output", default="output.xlsx", help="Output .xlsx (default: output.xlsx).")
    args = parser.parse_args()

    try:
        wb = load_workbook(args.input, data_only=True)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    tb_sheet_name = pick_tb_sheet(wb)
    mapping_sheet_name = pick_sheet(wb, MAPPING_SHEET_CANDIDATES)
    param_sheet_name = pick_sheet(wb, PARAMETERS_SHEET_CANDIDATES)

    try:
        accounts = read_tb(wb[tb_sheet_name])
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if not accounts:
        print("ERROR: No usable rows found in TB sheet.", file=sys.stderr)
        return 1

    tb_balances, _ = build_tb_dict(accounts)
    params = read_parameters(wb, param_sheet_name)

    tolerance = params.get("tolerance")
    tolerance = float(tolerance) if tolerance is not None else 0.01

    checks = []
    has_error = False

    def add_check(severity, message):
        nonlocal has_error
        checks.append((severity, message))
        if severity == "ERROR":
            has_error = True

    tb_sum = sum(acc["end_balance"] for acc in accounts if acc.get("end_balance") is not None)
    if abs(tb_sum) > tolerance:
        add_check("WARN", f"Trial balance not zero. Sum: {tb_sum:.2f}")

    derived_cash_begin, derived_cash_end = compute_cash_totals(accounts)
    cash_begin = params.get("cash_begin")
    cash_end = params.get("cash_end")
    cash_begin = float(cash_begin) if cash_begin is not None else derived_cash_begin
    cash_end = float(cash_end) if cash_end is not None else derived_cash_end

    wb_out = Workbook()
    wb_out.remove(wb_out.active)

    if mapping_sheet_name:
        try:
            mapping_rows = read_mapping(wb[mapping_sheet_name])
        except Exception as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1

        (
            section_order,
            line_order,
            line_totals,
            section_totals,
            used_accounts,
            missing_accounts,
        ) = apply_mapping(tb_balances, mapping_rows)

        if missing_accounts:
            add_check("ERROR", "Missing account_code(s): " + ", ".join(sorted(missing_accounts)))

        unmapped = set(tb_balances.keys()) - used_accounts
        if unmapped:
            add_check("WARN", "Unmapped account_code(s): " + ", ".join(sorted(unmapped)))

        assets, has_assets = sum_sections(
            section_totals.get("BS", {}),
            lambda s: section_match(s, ["asset", "资产"]),
        )
        liabilities, has_liab = sum_sections(
            section_totals.get("BS", {}),
            lambda s: section_match(s, ["liabilit", "负债"]),
        )
        equity, has_equity = sum_sections(
            section_totals.get("BS", {}),
            lambda s: section_match(s, ["equity", "权益", "capital", "所有者权益"]),
        )

        if not (has_assets and has_liab and has_equity):
            add_check("ERROR", "BS sections Assets/Liabilities/Equity not found.")
        else:
            diff = assets - (liabilities + equity)
            if abs(diff) > tolerance:
                add_check("ERROR", f"BS not balanced. Difference: {diff:.2f}")

        is_total = sum(section_totals.get("IS", {}).values())
        revenue, has_rev = sum_sections(
            section_totals.get("IS", {}),
            lambda s: section_match(s, ["revenue", "income", "收入"]),
        )
        expenses, has_exp = sum_sections(
            section_totals.get("IS", {}),
            lambda s: section_match(s, ["expense", "cost", "费用", "成本", "税金", "损失"]),
        )
        if has_rev or has_exp:
            net_profit = revenue - expenses
        else:
            net_profit = is_total
            add_check("WARN", "IS revenue/expense sections not identified; net profit uses total.")

        cf_total = sum(section_totals.get("CF", {}).values())
        if cash_begin is not None and cash_end is not None:
            diff = (cash_end - cash_begin) - cf_total
            if abs(diff) > tolerance:
                add_check("ERROR", f"CF net change mismatch. Difference: {diff:.2f}")
        else:
            add_check("WARN", "cash_begin/cash_end not provided; CF check skipped.")

        def write_statement(statement_key):
            ws = wb_out.create_sheet(statement_key)
            ws.append(["Section", "Line Item", "Amount"])
            for section in section_order.get(statement_key, []):
                for line_item in line_order[statement_key][section]:
                    amount = line_totals[statement_key][section][line_item]
                    ws.append([section, line_item, amount])
                ws.append([section, "TOTAL", section_totals[statement_key][section]])
            return ws

        write_statement("BS")
        ws_is = write_statement("IS")
        ws_is.append(["Profit", "NetProfit", net_profit])
        ws_cf = write_statement("CF")
        ws_cf.append(["Summary", "NetChangeInCash", cf_total])
    else:
        bs_entries = defaultdict(list)
        is_entries = defaultdict(list)
        unclassified = []

        for acc in accounts:
            end_balance = acc.get("end_balance")
            if end_balance is None:
                continue
            bs_section = classify_bs(acc.get("type"), acc.get("code", ""))
            if bs_section:
                bs_entries[bs_section].append(acc)
                continue
            is_section = classify_is(acc.get("type"), acc.get("name"), acc.get("code", ""))
            if is_section:
                is_entries[is_section].append(acc)
            else:
                unclassified.append(acc)

        assets_signed = sum(a["end_balance"] for a in bs_entries.get("Assets", []))
        liabilities_signed = sum(a["end_balance"] for a in bs_entries.get("Liabilities", []))
        equity_signed = sum(a["end_balance"] for a in bs_entries.get("Equity", []))

        if not (bs_entries.get("Assets") and bs_entries.get("Liabilities") and bs_entries.get("Equity")):
            add_check("ERROR", "BS sections Assets/Liabilities/Equity not identified in auto mode.")
        else:
            diff = assets_signed + liabilities_signed + equity_signed
            if abs(diff) > tolerance:
                add_check("ERROR", f"BS not balanced. Difference: {diff:.2f}")

        revenue_total = sum(abs(a["end_balance"]) for a in is_entries.get("Revenue", []))
        expense_total = sum(abs(a["end_balance"]) for a in is_entries.get("Expense", []))
        net_profit = revenue_total - expense_total
        if not is_entries.get("Revenue") and not is_entries.get("Expense"):
            add_check("WARN", "IS revenue/expense accounts not identified in auto mode.")

        if unclassified:
            sample = ", ".join(a.get("code", "") for a in unclassified[:10])
            add_check("WARN", f"Unclassified accounts: {sample}...")

        if cash_begin is None or cash_end is None:
            add_check("WARN", "cash_begin/cash_end not identified; CF summary limited.")

        ws_bs = wb_out.create_sheet("BS")
        ws_bs.append(["Section", "AccountCode", "AccountName", "Amount"])
        for section in ("Assets", "Liabilities", "Equity"):
            total = 0.0
            for acc in bs_entries.get(section, []):
                balance = acc["end_balance"]
                amount = balance if section == "Assets" else abs(balance)
                total += amount
                ws_bs.append([section, acc.get("code"), acc.get("name"), amount])
            ws_bs.append([section, "TOTAL", "", total])

        ws_is = wb_out.create_sheet("IS")
        ws_is.append(["Section", "AccountCode", "AccountName", "Amount"])
        for section in ("Revenue", "Expense"):
            total = 0.0
            for acc in is_entries.get(section, []):
                amount = abs(acc["end_balance"])
                total += amount
                ws_is.append([section, acc.get("code"), acc.get("name"), amount])
            ws_is.append([section, "TOTAL", "", total])
        ws_is.append(["Profit", "NetProfit", "", net_profit])

        ws_cf = wb_out.create_sheet("CF")
        ws_cf.append(["Item", "Amount"])
        if cash_begin is not None:
            ws_cf.append(["CashBegin", cash_begin])
        if cash_end is not None:
            ws_cf.append(["CashEnd", cash_end])
        if cash_begin is not None and cash_end is not None:
            ws_cf.append(["NetChangeInCash", cash_end - cash_begin])
        ws_cf.append(["NetProfit", net_profit])

        ws_unclassified = wb_out.create_sheet("Unclassified")
        ws_unclassified.append(["AccountCode", "AccountName", "EndBalance"])
        for acc in unclassified:
            ws_unclassified.append([acc.get("code"), acc.get("name"), acc.get("end_balance")])

    ws_checks = wb_out.create_sheet("Checks")
    ws_checks.append(["severity", "message"])
    for severity, message in checks:
        ws_checks.append([severity, message])

    wb_out.save(args.output)
    print(f"Saved output: {args.output}")
    return 1 if has_error else 0


if __name__ == "__main__":
    raise SystemExit(main())
