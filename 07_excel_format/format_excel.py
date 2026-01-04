import argparse
import sys
from datetime import date, datetime

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ACCOUNTING_FORMAT_DEC = "#,##0.00_);[Red](#,##0.00)"
ACCOUNTING_FORMAT_INT = "#,##0_);[Red](#,##0)"
DATE_FORMAT = "yyyy-mm-dd"


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def detect_columns(ws, header_row, scan_rows):
    numeric_cols = set()
    int_cols = set()
    date_cols = set()
    header_map = {}

    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(row=header_row, column=col).value
        header_map[col] = normalize_header(header_val)

        has_number = False
        has_decimal = False
        has_text = False
        is_date = "date" in header_map[col]
        max_row = min(ws.max_row, header_row + scan_rows)
        for row in range(header_row + 1, max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            if isinstance(value, (datetime, date)):
                is_date = True
                continue
            if isinstance(value, (int, float)):
                has_number = True
                if isinstance(value, float) and not value.is_integer():
                    has_decimal = True
                continue
            has_text = True

        if is_date:
            date_cols.add(col)
        elif has_number and not has_text:
            numeric_cols.add(col)
            if not has_decimal:
                int_cols.add(col)

    return numeric_cols, int_cols, date_cols


def set_column_widths(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            if isinstance(value, (datetime, date)):
                text = value.strftime("%Y-%m-%d")
            else:
                text = str(value)
            max_len = max(max_len, len(text))
        width = min(max(max_len + 2, 8), 40)
        ws.column_dimensions[get_column_letter(col)].width = width


def format_sheet(ws, header_row, scan_rows):
    numeric_cols, int_cols, date_cols = detect_columns(ws, header_row, scan_rows)

    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            if cell.column in date_cols:
                cell.number_format = DATE_FORMAT
            elif cell.column in int_cols:
                cell.number_format = ACCOUNTING_FORMAT_INT
            elif cell.column in numeric_cols:
                cell.number_format = ACCOUNTING_FORMAT_DEC

    set_column_widths(ws)


def main():
    parser = argparse.ArgumentParser(description="Apply financial-style formatting to Excel.")
    parser.add_argument("--input", default="input.xlsx", help="Input .xlsx (default: input.xlsx).")
    parser.add_argument("--output", default="output.xlsx", help="Output .xlsx (default: output.xlsx).")
    parser.add_argument("--sheets", default="", help="Comma-separated sheet names.")
    parser.add_argument("--header-row", type=int, default=1, help="Header row (default: 1).")
    parser.add_argument("--scan-rows", type=int, default=20, help="Rows to scan for types.")
    args = parser.parse_args()

    try:
        wb = load_workbook(args.input)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if args.sheets:
        sheet_names = [s.strip() for s in args.sheets.split(",") if s.strip()]
    else:
        sheet_names = wb.sheetnames

    for name in sheet_names:
        if name not in wb.sheetnames:
            print(f"ERROR: Sheet not found: {name}", file=sys.stderr)
            return 1
        format_sheet(wb[name], args.header_row, args.scan_rows)

    wb.save(args.output)
    print(f"Saved output: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
