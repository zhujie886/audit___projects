import argparse
import re
import sys
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook


HEADER_ALIASES = {
    "old_name": ["old_name", "old", "原文件名", "旧文件名", "原名称", "旧名称"],
    "new_name": ["new_name", "new", "新文件名", "新名称"],
}


def normalize_header(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[()（）\[\]【】:%/\\-]", "", text)
    return text


def find_header_index(headers, aliases):
    alias_norms = [normalize_header(a) for a in aliases]
    for alias in alias_norms:
        if alias in headers:
            return headers.index(alias)
    for idx, header in enumerate(headers):
        for alias in alias_norms:
            if alias and alias in header:
                return idx
    return None


def read_mapping(xlsx_path, sheet_name):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_header(value) for value in header_row]

    old_idx = find_header_index(headers, HEADER_ALIASES["old_name"])
    new_idx = find_header_index(headers, HEADER_ALIASES["new_name"])
    if old_idx is None or new_idx is None:
        raise ValueError("Missing required columns: old_name/new_name (或 原文件名/新文件名)")

    mapping = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        old_val = row[old_idx]
        new_val = row[new_idx]
        if old_val is None and new_val is None:
            continue
        if not old_val or not new_val:
            raise ValueError(f"Row {row_idx} has empty old_name or new_name.")
        mapping.append((str(old_val).strip(), str(new_val).strip()))

    return mapping


def validate_plan(plan, overwrite):
    targets = []
    for src, dst in plan:
        if src.resolve() == dst.resolve():
            continue
        if not src.exists():
            raise FileNotFoundError(f"Source not found: {src}")
        if src.is_dir():
            raise IsADirectoryError(f"Source is a directory, not a file: {src}")
        if dst.exists() and not overwrite:
            raise FileExistsError(f"Target already exists: {dst}")
        if dst.exists() and overwrite and dst.is_dir():
            raise IsADirectoryError(f"Target is a directory: {dst}")
        targets.append(str(dst.resolve()))

    duplicates = set()
    seen = set()
    for t in targets:
        if t in seen:
            duplicates.add(t)
        seen.add(t)
    if duplicates:
        raise ValueError("Duplicate target names: " + ", ".join(sorted(duplicates)))


def rename_files(plan, overwrite):
    temp_map = []
    # Two-phase rename avoids collisions (e.g., a->b and b->a).
    for src, dst in plan:
        if src.resolve() == dst.resolve():
            continue
        temp = src.with_name(src.name + ".renametmp_" + uuid4().hex)
        src.rename(temp)
        temp_map.append((temp, dst))

    for temp, dst in temp_map:
        if dst.exists() and overwrite:
            if dst.is_file():
                dst.unlink()
            else:
                raise IsADirectoryError(f"Target is a directory: {dst}")
        temp.rename(dst)


def main():
    parser = argparse.ArgumentParser(description="Rename files from an Excel mapping list.")
    parser.add_argument("--input", default="input.xlsx", help="Mapping .xlsx file (default: input.xlsx).")
    parser.add_argument("--folder", default="files", help="Target folder (default: files).")
    parser.add_argument("--sheet", default="", help="Sheet name (default: first sheet).")
    parser.add_argument("--dry-run", action="store_true", help="Print changes without renaming.")
    parser.add_argument("--overwrite", action="store_true", help="Allow overwriting existing files.")
    args = parser.parse_args()

    try:
        mapping = read_mapping(args.input, args.sheet or None)
        folder = Path(args.folder)
        if not folder.exists():
            raise FileNotFoundError(f"Folder not found: {folder}")
        plan = [(folder / old, folder / new) for old, new in mapping]
        validate_plan(plan, args.overwrite)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    if args.dry_run:
        for src, dst in plan:
            if src.resolve() == dst.resolve():
                continue
            print(f"{src} -> {dst}")
        return 0

    try:
        rename_files(plan, args.overwrite)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print(f"Renamed {len(plan)} file(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
